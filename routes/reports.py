from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook

from database import get_db
from models import ExpenseModel, IncomeModel


def auto_adjust_column_width(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        sheet.column_dimensions[column_letter].width = max_length + 3

router = APIRouter()


@router.get("/reports/excel")
def export_excel_report(db: Session = Depends(get_db)):
    incomes = db.query(IncomeModel).all()
    expenses = db.query(ExpenseModel).all()

    workbook = Workbook()

    income_sheet = workbook.active
    income_sheet.title = "Incomes"

    income_sheet.append(["ID", "Title", "Amount", "Source", "Date"])

    for income in incomes:
        income_sheet.append([
            income.id,
            income.title,
            income.amount,
            income.source,
            income.date
        ])

    expense_sheet = workbook.create_sheet(title="Expenses")

    expense_sheet.append(["ID", "Title", "Amount", "Category", "Date"])

    for expense in expenses:
        expense_sheet.append([
            expense.id,
            expense.title,
            expense.amount,
            expense.category,
            expense.date
        ])

    summary_sheet = workbook.create_sheet(title="Summary")

    total_income = 0
    total_expense = 0

    for income in incomes:
        total_income += income.amount

    for expense in expenses:
        total_expense += expense.amount

    balance = total_income - total_expense

    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Total Income", total_income])
    summary_sheet.append(["Total Expense", total_expense])
    summary_sheet.append(["Balance", balance])

    file_path = "budgetflow_report.xlsx"

    for row in income_sheet.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"

    for row in expense_sheet.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"

    auto_adjust_column_width(income_sheet)
    auto_adjust_column_width(expense_sheet)
    auto_adjust_column_width(summary_sheet)

    workbook.save(file_path)

    return FileResponse(
        path=file_path,
        filename="budgetflow_report.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )